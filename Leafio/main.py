import pandas as pd
import pyodbc
import ftplib
import schedule
import time
import os
from pathlib import Path
from datetime import datetime, timedelta
import zipfile
import shutil
import paramiko
from openpyxl import load_workbook

# Define database connection parameters
server = 'HSVSGNEDB07'
database = 'LSReport'
username = 'dd'
password = 'Hoahuongduong2908'
driver = 'SQL Server'

conn = pyodbc.connect(f'DRIVER={driver};SERVER={server};DATABASE={database};UID={username};PWD={password}', autocommit=True)

# create cursor object
cursor = conn.cursor()

cursor.execute("SELECT * FROM RunSQLtoCSV")
tab = cursor.fetchall()

cursor.execute("SELECT [Company], [No_] FROM LSReport.dbo.[LSC Store] WHERE [Store Type] = 0 AND [Attrib 1 Code] = 'Open' AND LEFT([No_],1) IN ('B','C') AND RIGHT([No_],3) NOT IN ('PRV','O2O','COM','HSV')")
store = cursor.fetchall()


def ftp_upload(path,ftp_host,ftp_port,ftp_user,ftp_pass,ftp_folder):
    try:
        # Connect to the FTP server
        ftp = ftplib.FTP()
        ftp.connect(host=ftp_host, port=ftp_port)
        ftp.login(user=ftp_user, passwd=ftp_pass)
        # if ftp_folder != "":
        #     ftp.cwd(ftp_folder)

        # Open the file in binary mode for reading
        with open(path, 'rb') as file:
            # Upload the file to the FTP server
            ftp.storbinary(f'STOR {os.path.basename(path)}', file)

        # Close the FTP connection
        ftp.quit()

        os.remove(path)
    except Exception as e:
        print(e) 

def sftp_upload(path,ftp_host,ftp_port,ftp_user,ftp_pass,ftp_folder):
    try:
        # Connect to the FTP server
        ssh = paramiko.SSHClient()
        ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        ssh.connect(hostname=ftp_host, port=ftp_port, username=ftp_user, password=ftp_pass)
        # Open an SFTP session
        sftp = ssh.open_sftp()
        if ftp_folder != "":
            sftp.chdir(ftp_folder)
        # Extract the filename from the local CSV path
            csv_filename = path.rsplit('\\', 1)[-1]
        # Upload the local CSV file to the SFTP server
            sftp.put(path, csv_filename)

        # Close the FTP connection
        sftp.close()
        ssh.close()

        os.remove(path)
    except Exception as e:
        print(e) 


def sql_execute_toexcel(sql_query,template_path,output_path,sheet_index):
    try:
        if sheet_index == 0:
            # Sao chép tệp Excel template
            shutil.copy(template_path, output_path)
        with pyodbc.connect(f'DRIVER={driver};SERVER={server};DATABASE={database};UID={username};PWD={password}', autocommit=True) as cnxn:
            df = pd.read_sql_query(sql_query, cnxn)
            if not df.empty:
                # Load the existing Excel file
                book = load_workbook(output_path)
                writer = pd.ExcelWriter(output_path, engine='openpyxl') 
                writer.book = book

                # Append the DataFrame to the existing sheet
                sheet_names = book.sheetnames
                sheet_name = sheet_names[sheet_index]
                df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=writer.sheets[sheet_name].max_row, header=False)

                # Save the changes
                writer.save()

            # Clear the DataFrame when done
            df.drop(index=df.index, inplace=True)
            df.drop(columns=df.columns, inplace=True)
        cnxn.close()
    except Exception as e:
        print(e)        

def compress_folder(folder_path, zip_path):
    # Create a zip file object
    with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
        # Iterate over all files and subdirectories in the folder
        for root, _, files in os.walk(folder_path):
            for file in files:
                file_path = os.path.join(root, file)
                # Add the file to the zip file
                zipf.write(file_path, arcname=os.path.relpath(file_path, folder_path))



def sql_execute(sql_query,csv_path):
    # Connect to database using connection pooling
    try:
        with pyodbc.connect(f'DRIVER={driver};SERVER={server};DATABASE={database};UID={username};PWD={password}', autocommit=True) as cnxn:
            df = pd.read_sql_query(sql_query, cnxn)  
            if not df.empty:   
                # Save the DataFrame to a CSV file
                df.to_csv(csv_path, index=False)
                # clear the DataFrame when done
                df.drop(index=df.index, inplace=True)  # drops all rows
                df.drop(columns=df.columns, inplace=True)  # drops all columns      
        cnxn.close() 
    except Exception as e:
        print(e)  

def delete_localfolder(folder_path):
    if os.path.exists(folder_path):
        shutil.rmtree(folder_path)

def leafio():
    #Check folder local
    leafio = 0
    folder_path = 'C:\HSV.Services\Leafio'
    folder_name = f"{(datetime.now()- timedelta(days=1)).strftime('%Y%m%d%H%M')}"
    folder_path = os.path.join(folder_path, folder_name)
    if os.path.exists(folder_path) == False:
        os.makedirs(folder_path)

    for row in tab:
        if row[13] == "Leafio":
            leafio = 1
            ftp_host = row[2]
            ftp_port = row[3]
            ftp_user = row[4]
            ftp_pass = row[5]
            ftp_folder = row[6]
            if row[12] == True and datetime.strptime(row[10], '%Y-%m-%d').date() <= datetime.now().date() and datetime.strptime(row[11], '%Y-%m-%d').date() >= datetime.now().date() and row[9] == False:
               # define sql query
                sql_query = "EXEC " + row[1]
                # Define file paths
                csv_path = os.path.join(folder_path, Path(f"{row[7]}{(datetime.now()- timedelta(days=1)).strftime('%Y%m%d')}.csv"))
                sql_execute(sql_query,csv_path)
            else :
                if row[12] == True and datetime.strptime(row[10], '%Y-%m-%d').date() <= datetime.now().date() and datetime.strptime(row[11], '%Y-%m-%d').date() >= datetime.now().date() and row[9] == True:
                    for i in store:
                        if i[0] == row[8]:
                            # define sql query
                            sql_query = f"EXEC {row[1]} @store = '{i[1]}'"
                            # Define file paths
                            csv_path = os.path.join(folder_path, Path(f"{row[7]}{i[1]}_{(datetime.now()- timedelta(days=1)).strftime('%Y%m%d')}.csv"))
                            sql_execute(sql_query,csv_path)
    if leafio == 1:
        zip_path = os.path.join('C:\HSV.Services\Leafio', Path(f"{(datetime.now()- timedelta(days=1)).strftime('%Y%m%d%H%M')}.zip"))
        compress_folder(folder_path, zip_path)
        ftp_upload(zip_path,ftp_host,ftp_port,ftp_user,ftp_pass,ftp_folder)
        delete_localfolder(folder_path)
        
def palexy():
    palexy = 0
    folder_path = 'C:\HSV.Services\Leafio\Palexy'
    folder_name = f"{(datetime.now()- timedelta(days=1)).strftime('%Y%m%d%H%M')}"
    folder_path = os.path.join(folder_path, folder_name)
    if os.path.exists(folder_path) == False:
        os.makedirs(folder_path)
    
    for row in tab:
        if row[13] == "Palexy":
            palexy = 1
            ftp_host = row[2]
            ftp_port = row[3]
            ftp_user = row[4]
            ftp_pass = row[5]
            ftp_folder = row[6]      
            if row[12] == True and datetime.strptime(row[10], '%Y-%m-%d').date() <= datetime.now().date() and datetime.strptime(row[11], '%Y-%m-%d').date() >= datetime.now().date():
                  # define sql query
                   sql_query = "EXEC " + row[1]
                   # Define file paths
                   csv_path = os.path.join(folder_path, Path(f"{row[7]}{(datetime.now()- timedelta(days=1)).strftime('%Y%m%d')}.csv"))
                   sql_execute(sql_query,csv_path)
    if palexy == 1:
        sftp_upload(csv_path,ftp_host,ftp_port,ftp_user,ftp_pass,ftp_folder)
        delete_localfolder(folder_path)

def nielsen():
    nielsen = 0
    folder_path = 'C:/HSV.Services/Leafio/Nielsen'
    if os.path.exists(folder_path) == False:
        os.makedirs(folder_path)
    
    for row in tab:
        if row[13] == "Nielsen":
            nielsen = 1
            ftp_host = row[2]
            ftp_port = row[3]
            ftp_user = row[4]
            ftp_pass = row[5]
            ftp_folder = row[6]      
            if row[12] == True and datetime.strptime(row[10], '%Y-%m-%d').date() <= datetime.now().date() and datetime.strptime(row[11], '%Y-%m-%d').date() >= datetime.now().date():
                   for step in range(0,3):
                        # define sql query
                         sql_query = "EXEC " + row[1] + " @step = " + str(step)
                         # Define file paths
                         template_path = 'C:/HSV.Services/Leafio/Nielsen/template/ITEMSALESBYSTORE_TEMPLATE.xlsx'
                         output_path = os.path.join(folder_path, Path(f"{row[7]}{(datetime.now()- timedelta(days=1)).strftime('%Y_%m_%d')}.xlsx"))
                         sql_execute_toexcel(sql_query,template_path,output_path,step)
    if nielsen == 1:
        sftp_upload(output_path,ftp_host,ftp_port,ftp_user,ftp_pass,ftp_folder)

# close the cursor and the connection
cursor.close()
conn.close()

# Lập lịch chạy
schedule.every().day.at("05:00").do(leafio)
schedule.every().day.at("06:00").do(palexy)
schedule.every().monday.at("06:30").do(nielsen)

while True:
    schedule.run_pending()
    time.sleep(1)

